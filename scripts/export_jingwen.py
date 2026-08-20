#!/usr/bin/env python3
"""把 App 内 Hexagrams.json 导出成可编辑的 Excel（易经正文编辑表.xlsx）。"""
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
HEXES = ROOT / "ios/Yizhidao/Resources/Hexagrams.json"
OUT = ROOT / "易经正文编辑表.xlsx"

FONT = Font(name="Arial", size=11)
FONT_BOLD = Font(name="Arial", size=11, bold=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True)
ID_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E8E0D4", end_color="E8E0D4", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

YAO_LABELS = ["初爻", "二爻", "三爻", "四爻", "五爻", "上爻"]

raw = json.loads(HEXES.read_text(encoding="utf-8"))
hexes = raw["hexagrams"] if isinstance(raw, dict) else raw
wings = raw.get("wings", []) if isinstance(raw, dict) else []
hexes = sorted(hexes, key=lambda h: h["number"])


def style_header(ws):
    for cell in ws[1]:
        cell.font = FONT_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def style_body(ws, id_cols, widths, row_height=48):
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col):
        for i, cell in enumerate(row, start=1):
            cell.font = FONT
            cell.alignment = WRAP
            if i in id_cols:
                cell.fill = ID_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = row_height
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{ws.max_row}"


wb = Workbook()
ws_info = wb.active
ws_info.title = "说明"
info = [
    ("《易经》正文编辑表", FONT_TITLE),
    ("", FONT),
    ("对应 App 内 ios/Yizhidao/Resources/Hexagrams.json（六十四卦 + 文言 + 四传）。", FONT),
    ("审查：在各表用筛选（卦名 / 爻位 / 传）。", FONT),
    ("修改：直接改单元格后保存。灰底列（卦号、爻位、段序）请勿改。", FONT),
    ("「六十四卦」：一卦一行。卦辞、彖、大象、用九/用六。无用九用六的格子留空。", FONT),
    ("「爻辞」：一爻一行，初→上。爻辞与小象成对。", FONT),
    ("「文言」：乾、坤分段。可增删行（段序按数字排序写回）。", FONT),
    ("「四传」：系辞 / 说卦 / 序卦 / 杂卦。可改正文；传、章名请与原表一致。", FONT),
    ("改完保存本文件，在对话里说「写回经文」，会同步到 Hexagrams.json。", FONT),
]
for i, (text, font) in enumerate(info, start=1):
    cell = ws_info.cell(i, 1, text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
ws_info.column_dimensions["A"].width = 96
ws_info.row_dimensions[1].height = 22
for r in range(3, 12):
    ws_info.row_dimensions[r].height = 22

ws_g = wb.create_sheet("六十四卦")
g_headers = ["卦号", "卦名", "上下经", "目录标题", "卦象题", "卦辞", "彖", "大象", "用辞", "用象"]
ws_g.append(g_headers)
for h in hexes:
    yong = h.get("yong") or {}
    ws_g.append([
        h["number"],
        h.get("name", ""),
        h.get("part", ""),
        h.get("title", ""),
        h.get("figure", ""),
        h.get("guaci", ""),
        h.get("tuanci", ""),
        h.get("daxiang", ""),
        yong.get("ci", "") if isinstance(yong, dict) else "",
        yong.get("xiang", "") if isinstance(yong, dict) else "",
    ])
style_header(ws_g)
style_body(ws_g, id_cols={1}, widths=[8, 10, 10, 18, 22, 36, 48, 36, 28, 28], row_height=56)

ws_y = wb.create_sheet("爻辞")
ws_y.append(["卦号", "卦名", "爻位", "爻辞", "小象"])
for h in hexes:
    yaoci = h.get("yaoci") or [""] * 6
    xiaoxiang = h.get("xiaoxiang") or [""] * 6
    for i, label in enumerate(YAO_LABELS):
        ws_y.append([
            h["number"],
            h.get("name", ""),
            label,
            yaoci[i] if i < len(yaoci) else "",
            xiaoxiang[i] if i < len(xiaoxiang) else "",
        ])
style_header(ws_y)
style_body(ws_y, id_cols={1, 3}, widths=[8, 10, 10, 40, 40], row_height=36)

ws_w = wb.create_sheet("文言")
ws_w.append(["卦号", "卦名", "段序", "正文"])
for h in hexes:
    for i, para in enumerate(h.get("wenyan") or [], start=1):
        ws_w.append([h["number"], h.get("name", ""), i, para])
style_header(ws_w)
style_body(ws_w, id_cols={1, 3}, widths=[8, 10, 8, 80], row_height=56)

ws_t = wb.create_sheet("四传")
ws_t.append(["传", "章", "段序", "正文"])
for wing in wings:
    wing_title = wing.get("title", "")
    for chapter in wing.get("chapters") or []:
        chap_title = chapter.get("title", "")
        for i, para in enumerate(chapter.get("paragraphs") or [], start=1):
            ws_t.append([wing_title, chap_title, i, para])
style_header(ws_t)
style_body(ws_t, id_cols={3}, widths=[12, 14, 8, 88], row_height=56)

wb.save(OUT)
print(f"已导出 {len(hexes)} 卦 / {sum(len(h.get('wenyan') or []) for h in hexes)} 段文言 / "
      f"{sum(len(ch.get('paragraphs') or []) for w in wings for ch in w.get('chapters') or [])} 段四传 -> {OUT}")
