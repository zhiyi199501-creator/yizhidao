#!/usr/bin/env python3
"""把 App 内 cases.json 导出成可编辑的 Excel（案例编辑表.xlsx）。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import json

from collections import Counter

ROOT = Path(__file__).resolve().parents[1]
CASES = ROOT / "Yizhidao/Resources/cases.json"
HEXES = ROOT / "Yizhidao/Resources/Hexagrams.json"
OUT = ROOT / "案例编辑表.xlsx"

FONT = Font(name="Arial", size=11)
FONT_BOLD = Font(name="Arial", size=11, bold=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True)
ID_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E8E0D4", end_color="E8E0D4", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

cases = json.loads(CASES.read_text(encoding="utf-8"))
_raw_hexes = json.loads(HEXES.read_text(encoding="utf-8"))
hexes = _raw_hexes["hexagrams"] if isinstance(_raw_hexes, dict) else _raw_hexes
cases.sort(key=lambda c: (c.get("number") or 99, c.get("file") or ""))

headers = ["编号", "卦名", "爻位", "背景", "所问何事", "起卦结果", "讲师解读", "验证结果"]
fields = ["file", "hexagram", "position", "background", "question", "casting", "explanation", "verification"]

wb = Workbook()

ws_info = wb.active
ws_info.title = "说明"
info = [
    ("《张庆祥讲易经案例》编辑表", FONT_TITLE),
    ("", FONT),
    ("审查：在「案例」表用筛选（卦名 / 爻位）；「按卦」表看每卦条数（缺谦卦）。", FONT),
    ("修改：直接改单元格后保存。编号列（灰底）请勿改；改了会被当成新案例。", FONT),
    ("新增：表格末尾加一行，编号留空。", FONT),
    ("删除：删整行。", FONT),
    ("卦名：写「屯」或「屯卦」均可，写回 App 时会重新对应卦号。", FONT),
    ("爻位：初爻 / 二爻 / 三爻 / 四爻 / 五爻 / 上爻 / 卦辞（可写「三爻、四爻」）。", FONT),
    ("改完保存本文件，在对话里说「写回案例」，会同步到 Yizhidao/Resources/cases.json。", FONT),
]
for i, (text, font) in enumerate(info, start=1):
    cell = ws_info.cell(i, 1, text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
ws_info.column_dimensions["A"].width = 92
ws_info.row_dimensions[1].height = 22
for r in range(3, 10):
    ws_info.row_dimensions[r].height = 22

ws = wb.create_sheet("案例")
ws.append(headers)
for c in cases:
    ws.append([c.get(f, "") for f in fields])

for cell in ws[1]:
    cell.font = FONT_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
    for i, cell in enumerate(row):
        cell.font = FONT
        cell.alignment = WRAP
        if i == 0:
            cell.fill = ID_FILL

widths = [40, 12, 12, 48, 40, 28, 56, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 48
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"

ws_g = wb.create_sheet("按卦")
for col, title in enumerate(["卦号", "卦名", "案例表卦名", "案例数"], start=1):
    cell = ws_g.cell(1, col, title)
    cell.font = FONT_BOLD
    cell.fill = HEADER_FILL
counts = Counter(c.get("hexagram") for c in cases)
total = 0
for i, h in enumerate(hexes, start=2):
    name = h["name"]
    full = name if name.endswith("卦") else name + "卦"
    n = counts.get(full, 0) + counts.get(name, 0)
    total += n
    ws_g.cell(i, 1, h["number"]).font = FONT
    ws_g.cell(i, 2, name).font = FONT
    ws_g.cell(i, 3, full).font = FONT
    cell = ws_g.cell(i, 4, n)
    cell.font = FONT
    cell.alignment = Alignment(horizontal="right")
    if n == 0:
        cell.fill = PatternFill(start_color="F8E6E6", end_color="F8E6E6", fill_type="solid")
ws_g.cell(66, 3, "合计").font = FONT_BOLD
ws_g.cell(66, 4, total).font = FONT_BOLD
ws_g.cell(66, 4).alignment = Alignment(horizontal="right")
note = ws_g.cell(68, 1, "案例数来自 Yizhidao/Resources/cases.json 导出当时的条数（2026-08-17）。改「案例」表后重新运行 scripts/export_cases.py 会刷新本表。")
note.font = Font(name="Arial", size=10, italic=True)
note.alignment = Alignment(wrap_text=True)
ws_g.merge_cells("A68:D69")
ws_g.column_dimensions["A"].width = 10
ws_g.column_dimensions["B"].width = 12
ws_g.column_dimensions["C"].width = 16
ws_g.column_dimensions["D"].width = 12
ws_g.freeze_panes = "A2"
ws_g.auto_filter.ref = "A1:D65"

wb.save(OUT)
print(f"已导出 {len(cases)} 条 -> {OUT}")
